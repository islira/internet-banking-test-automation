#!/usr/bin/env python3
"""
Automação para atualizar matriz de rastreabilidade com resultados dos testes Playwright
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TraceabilityMatrixUpdater:
    """Classe para atualizar a matriz de rastreabilidade com resultados dos testes"""
    
    # Cores para status dos testes
    COLORS = {
        'passed': '90EE90',    # Verde claro
        'failed': 'FFB6C1',    # Vermelho claro
        'skipped': 'FFD700',   # Amarelo
        'timedOut': 'FFA500',  # Laranja
        'interrupted': 'D3D3D3' # Cinza
    }
    
    def __init__(self, results_file: str = 'test-results.json', 
                 matrix_file: str = 'matriz_rastreabilidade.xlsx'):
        """
        Inicializa o atualizador
        
        Args:
            results_file: Caminho para o arquivo JSON com resultados dos testes
            matrix_file: Caminho para o arquivo Excel da matriz de rastreabilidade
        """
        self.results_file = Path(results_file)
        self.matrix_file = Path(matrix_file)
        self.results_data = None
        
    def load_test_results(self) -> Dict[str, Any]:
        """Carrega os resultados dos testes do arquivo JSON"""
        if not self.results_file.exists():
            raise FileNotFoundError(
                f"Arquivo de resultados não encontrado: {self.results_file}\n"
                "Execute os testes com: npx playwright test --reporter=json"
            )
        
        with open(self.results_file, 'r', encoding='utf-8') as f:
            self.results_data = json.load(f)
        
        return self.results_data
    
    def extract_test_info(self) -> List[Dict[str, Any]]:
        """Extrai informações relevantes dos testes"""
        if not self.results_data:
            self.load_test_results()
        
        tests_info = []
        
        for suite in self.results_data.get('suites', []):
            suite_title = suite.get('title', 'Unknown Suite')
            
            for spec in suite.get('specs', []):
                test_title = spec.get('title', 'Unknown Test')
                
                for test in spec.get('tests', []):
                    for result in test.get('results', []):
                        status = result.get('status', 'unknown')
                        duration = result.get('duration', 0) / 1000  # Converter para segundos
                        
                        # Extrair arquivo do teste
                        test_file = ''
                        if suite.get('file'):
                            test_file = Path(suite['file']).name
                        
                        tests_info.append({
                            'suite': suite_title,
                            'test_file': test_file,
                            'test_name': test_title,
                            'status': status,
                            'duration': round(duration, 2),
                            'error': result.get('error', {}).get('message', '') if status == 'failed' else '',
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        })
        
        return tests_info
    
    def create_matrix_if_not_exists(self):
        """Cria uma nova matriz de rastreabilidade se não existir"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz de Rastreabilidade"
        
        # Definir cabeçalhos
        headers = [
            'ID',
            'Suite de Teste',
            'Arquivo',
            'Nome do Teste',
            'Status',
            'Duração (s)',
            'Última Execução',
            'Erro/Observação'
        ]
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar largura das colunas
        column_widths = [8, 30, 20, 50, 12, 12, 20, 40]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        wb.save(self.matrix_file)
        print(f"✓ Matriz de rastreabilidade criada: {self.matrix_file}")
    
    def update_matrix(self, tests_info: List[Dict[str, Any]]):
        """Atualiza a matriz de rastreabilidade com os resultados dos testes"""
        if not self.matrix_file.exists():
            self.create_matrix_if_not_exists()
        
        wb = openpyxl.load_workbook(self.matrix_file)
        ws = wb.active
        
        # Criar um dicionário dos testes existentes (chave: suite + test_name)
        existing_tests = {}
        for row_num in range(2, ws.max_row + 1):
            suite = ws.cell(row=row_num, column=2).value
            test_name = ws.cell(row=row_num, column=4).value
            if suite and test_name:
                key = f"{suite}|{test_name}"
                existing_tests[key] = row_num
        
        # Preparar estilos
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Atualizar ou adicionar testes
        next_id = ws.max_row if ws.max_row > 1 else 1
        
        for test_info in tests_info:
            key = f"{test_info['suite']}|{test_info['test_name']}"
            
            if key in existing_tests:
                # Atualizar teste existente
                row_num = existing_tests[key]
            else:
                # Adicionar novo teste
                row_num = next_id + 1
                next_id += 1
            
            # Preencher dados
            ws.cell(row=row_num, column=1).value = row_num - 1  # ID
            ws.cell(row=row_num, column=2).value = test_info['suite']
            ws.cell(row=row_num, column=3).value = test_info['test_file']
            ws.cell(row=row_num, column=4).value = test_info['test_name']
            ws.cell(row=row_num, column=5).value = test_info['status'].upper()
            ws.cell(row=row_num, column=6).value = test_info['duration']
            ws.cell(row=row_num, column=7).value = test_info['timestamp']
            ws.cell(row=row_num, column=8).value = test_info['error']
            
            # Aplicar cores baseadas no status
            status_color = self.COLORS.get(test_info['status'], 'FFFFFF')
            status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            
            # Aplicar formatação
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = alignment
                
                # Aplicar cor de fundo na coluna de status
                if col_num == 5:
                    cell.fill = status_fill
                    cell.font = Font(bold=True)
        
        # Salvar arquivo
        wb.save(self.matrix_file)
        print(f"✓ Matriz de rastreabilidade atualizada: {self.matrix_file}")
    
    def generate_summary(self, tests_info: List[Dict[str, Any]]):
        """Gera um resumo dos resultados dos testes"""
        total = len(tests_info)
        passed = sum(1 for t in tests_info if t['status'] == 'passed')
        failed = sum(1 for t in tests_info if t['status'] == 'failed')
        skipped = sum(1 for t in tests_info if t['status'] == 'skipped')
        
        print("\n" + "="*60)
        print("RESUMO DA EXECUÇÃO DOS TESTES")
        print("="*60)
        print(f"Total de testes:    {total}")
        print(f"✓ Aprovados:        {passed} ({passed/total*100:.1f}%)")
        print(f"✗ Falhas:           {failed} ({failed/total*100:.1f}%)")
        print(f"○ Ignorados:        {skipped} ({skipped/total*100:.1f}%)")
        print(f"Data/Hora:          {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*60 + "\n")
    
    def run(self):
        """Executa o processo completo de atualização"""
        try:
            print("Iniciando atualização da matriz de rastreabilidade...")
            print(f"Arquivo de resultados: {self.results_file}")
            print(f"Matriz de rastreabilidade: {self.matrix_file}\n")
            
            # Carregar e processar resultados
            self.load_test_results()
            tests_info = self.extract_test_info()
            
            if not tests_info:
                print("⚠ Nenhum teste encontrado nos resultados.")
                return
            
            # Atualizar matriz
            self.update_matrix(tests_info)
            
            # Gerar resumo
            self.generate_summary(tests_info)
            
            print("✓ Processo concluído com sucesso!")
            
        except FileNotFoundError as e:
            print(f"✗ Erro: {e}")
        except Exception as e:
            print(f"✗ Erro inesperado: {e}")
            raise


def main():
    """Função principal"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Atualiza matriz de rastreabilidade com resultados do Playwright'
    )
    parser.add_argument(
        '--results',
        default='test-results.json',
        help='Caminho para o arquivo JSON de resultados (padrão: test-results.json)'
    )
    parser.add_argument(
        '--matrix',
        default='matriz_rastreabilidade.xlsx',
        help='Caminho para a matriz de rastreabilidade (padrão: matriz_rastreabilidade.xlsx)'
    )
    
    args = parser.parse_args()
    
    updater = TraceabilityMatrixUpdater(
        results_file=args.results,
        matrix_file=args.matrix
    )
    updater.run()


if __name__ == '__main__':
    main()
